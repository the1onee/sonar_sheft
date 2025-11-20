from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Q
from .models import EmployeeAssignment, Employee, Sonar, Shift, WeeklyShiftAssignment, Supervisor, AssignmentConfirmation, SystemSettings, Manager, CustomNotification
from .forms import EmployeeAssignmentForm, LoginForm, EmployeeForm, SonarForm, ShiftForm, WeeklyShiftAssignmentForm, SystemSettingsForm, ManagerCreateForm, SupervisorCreateForm, EmployeeAccountCreateForm, CustomNotificationForm
from .utils import send_telegram_message
from .decorators import get_user_role, superadmin_required, manager_required, supervisor_required, employee_required, staff_required

# صفحة الهبوط (Landing Page)
def landing_page(request):
    """صفحة الهبوط الرئيسية قبل تسجيل الدخول"""
    if request.user.is_authenticated:
        return redirect('home')
    return render(request, 'landing.html')

# دالة للتحقق من أن المستخدم مشرف (supervisor) وليس admin
def is_supervisor(user):
    """فحص إذا كان المستخدم مشرف وليس admin"""
    return hasattr(user, 'supervisor_profile') and user.supervisor_profile.is_active

# صفحة تسجيل الدخول
def user_login(request):
    if request.user.is_authenticated:
        # إذا كان مشرف، أرسله للـ home
        if is_supervisor(request.user):
            return redirect('home')
        return redirect('home')
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                
                # فحص نوع المستخدم
                user_type = "مشرف" if is_supervisor(user) else "مسؤول النظام"
                messages.success(request, f'مرحباً {user.first_name or user.username}! ({user_type}) تم تسجيل الدخول بنجاح 👋')
                
                # المشرف دائماً يذهب للـ home
                if is_supervisor(user):
                    return redirect('home')
                
                next_url = request.GET.get('next', 'home')
                return redirect(next_url)
            else:
                messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة!')
        else:
            messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة!')
    else:
        form = LoginForm()
    
    return render(request, 'login.html', {'form': form})

# تسجيل الخروج
@login_required
def user_logout(request):
    logout(request)
    messages.info(request, 'تم تسجيل الخروج بنجاح. نراك قريباً! 👋')
    return redirect('login')

# الصفحة الرئيسية (تتطلب تسجيل دخول)
@login_required
def home(request):
    """الصفحة الرئيسية - تتغير حسب دور المستخدم"""
    user_role = get_user_role(request.user)
    
    # إعادة توجيه حسب الدور
    if user_role == 'employee':
        return redirect('employee_dashboard')
    elif user_role == 'supervisor':
        return redirect('supervisor_dashboard')
    elif user_role == 'manager':
        return redirect('manager_dashboard')
    elif user_role == 'superadmin':
        return redirect('admin_dashboard')
    
    # في حالة عدم وجود دور محدد
    messages.warning(request, 'لم يتم تحديد دور لحسابك. يرجى التواصل مع المدير.')
    return redirect('login')


# ==================== Dashboards حسب الأدوار ====================

@superadmin_required
def admin_dashboard(request):
    """لوحة تحكم السوبر أدمن"""
    # إحصائيات شاملة
    total_managers = Manager.objects.filter(is_active=True).count()
    total_supervisors = Supervisor.objects.filter(is_active=True).count()
    total_employees = Employee.objects.count()
    total_sonars = Sonar.objects.count()
    total_assignments = EmployeeAssignment.objects.count()
    pending_assignments = EmployeeAssignment.objects.filter(employee_confirmed=False).count()
    
    # آخر العمليات
    recent_assignments = EmployeeAssignment.objects.order_by('-assigned_at')[:10]
    
    context = {
        'user_role': 'superadmin',
        'total_managers': total_managers,
        'total_supervisors': total_supervisors,
        'total_employees': total_employees,
        'total_sonars': total_sonars,
        'total_assignments': total_assignments,
        'pending_assignments': pending_assignments,
        'recent_assignments': recent_assignments,
    }
    return render(request, 'dashboards/admin.html', context)


@manager_required
def manager_dashboard(request):
    """لوحة تحكم المدير"""
    # إحصائيات المدير
    total_supervisors = Supervisor.objects.filter(created_by=request.user, is_active=True).count()
    total_employees = Employee.objects.filter(created_by=request.user).count()
    total_sonars = Sonar.objects.count()
    pending_assignments = EmployeeAssignment.objects.filter(employee_confirmed=False).count()
    
    # الإعدادات الحالية
    settings = SystemSettings.get_current_settings()
    
    # آخر العمليات
    recent_supervisors = Supervisor.objects.filter(created_by=request.user).order_by('-created_at')[:5]
    recent_employees = Employee.objects.filter(created_by=request.user).order_by('-created_at')[:5]
    
    context = {
        'user_role': 'manager',
        'total_supervisors': total_supervisors,
        'total_employees': total_employees,
        'total_sonars': total_sonars,
        'pending_assignments': pending_assignments,
        'settings': settings,
        'recent_supervisors': recent_supervisors,
        'recent_employees': recent_employees,
    }
    return render(request, 'dashboards/manager.html', context)


@supervisor_required
def supervisor_dashboard(request):
    """لوحة تحكم المشرف"""
    # إحصائيات المشرف
    pending_assignments = EmployeeAssignment.objects.filter(employee_confirmed=False).count()
    confirmed_today = EmployeeAssignment.objects.filter(
        supervisor_confirmed_at__date=timezone.now().date(),
        supervisor_confirmed=True,
        supervisor_confirmed_by=request.user
    ).count()
    employees_on_leave = Employee.objects.filter(is_on_leave=True).count()
    active_sonars = Sonar.objects.filter(active=True).count()
    inactive_sonars = Sonar.objects.filter(active=False).count()
    
    # 📊 إحصائيات الموظفين وساعات العمل
    employees_stats = []
    all_employees = Employee.objects.filter(is_on_leave=False).order_by('name')
    
    # حساب متوسط ساعات العمل
    total_work_hours = sum(emp.total_work_hours for emp in all_employees)
    avg_work_hours = total_work_hours / all_employees.count() if all_employees.count() > 0 else 0.0
    
    for emp in all_employees:
        diff_from_avg = emp.total_work_hours - avg_work_hours
        employees_stats.append({
            'employee': emp,
            'total_work_hours': emp.total_work_hours,
            'diff_from_avg': diff_from_avg,
            'last_work': emp.last_work_datetime,
            'consecutive_rest': emp.consecutive_rest_count,
            'priority_score': emp.get_priority_score(avg_work_hours)
        })
    
    # ترتيب حسب ساعات العمل (من الأكثر إلى الأقل)
    employees_stats.sort(key=lambda x: x['total_work_hours'], reverse=True)
    
    # تقسيم الموظفين إلى فئات
    top_workers = [e for e in employees_stats if e['diff_from_avg'] > 1.0][:5]  # أكثر 5 عملاً
    need_work = [e for e in employees_stats if e['diff_from_avg'] < -1.0][:5]  # أقل 5 عملاً
    
    # التبديلات المؤكدة من الموظفين (بانتظار تأكيد المشرف)
    waiting_supervisor_confirmation = EmployeeAssignment.objects.filter(
        employee_confirmed=True,
        supervisor_confirmed=False
    ).select_related('employee', 'sonar', 'shift').order_by('-employee_confirmed_at')
    
    # جميع التبديلات (غير مؤكدة، مؤكدة من الموظف، مؤكدة من المشرف)
    all_assignments = EmployeeAssignment.objects.select_related(
        'employee', 'sonar', 'shift', 'supervisor_confirmed_by'
    ).order_by('-assigned_at')[:50]  # آخر 50 تبديل
    
    # إحصائيات التأكيد
    waiting_employee_count = EmployeeAssignment.objects.filter(
        employee_confirmed=False
    ).count()
    
    waiting_supervisor_count = EmployeeAssignment.objects.filter(
        employee_confirmed=True,
        supervisor_confirmed=False
    ).count()
    
    fully_confirmed_count = EmployeeAssignment.objects.filter(
        supervisor_confirmed=True
    ).count()
    
    # التبديلات المعلقة
    pending_list = EmployeeAssignment.objects.filter(
        employee_confirmed=False
    ).select_related('employee', 'sonar', 'shift').order_by('-assigned_at')[:10]
    
    context = {
        'user_role': 'supervisor',
        'pending_assignments': pending_assignments,
        'confirmed_today': confirmed_today,
        'employees_on_leave': employees_on_leave,
        'active_sonars': active_sonars,
        'inactive_sonars': inactive_sonars,
        'pending_list': pending_list,
        'waiting_supervisor_confirmation': waiting_supervisor_confirmation,
        'all_assignments': all_assignments,
        'waiting_employee_count': waiting_employee_count,
        'waiting_supervisor_count': waiting_supervisor_count,
        'fully_confirmed_count': fully_confirmed_count,
        # إحصائيات الموظفين الجديدة
        'employees_stats': employees_stats,
        'avg_work_hours': avg_work_hours,
        'top_workers': top_workers,
        'need_work': need_work,
    }
    return render(request, 'dashboards/supervisor.html', context)


@employee_required
def employee_dashboard(request):
    """لوحة تحكم الموظف"""
    try:
        employee = request.user.employee_profile
    except:
        messages.error(request, 'لم يتم العثور على ملف الموظف الخاص بك!')
        return redirect('login')
    
    # إحصائيات الموظف
    today = timezone.now().date()
    
    # التبديلات اليوم
    today_assignments = EmployeeAssignment.objects.filter(
        employee=employee,
        assigned_at__date=today
    ).select_related('sonar', 'shift').order_by('assigned_at')
    
    # عدد السونارات التي ذهب إليها اليوم
    sonars_today = today_assignments.values('sonar').distinct().count()
    
    # عدد التبديلات اليوم
    total_shifts_today = today_assignments.count()
    
    # التبديلات المؤكدة اليوم
    confirmed_today = today_assignments.filter(supervisor_confirmed=True).count()
    
    # التبديلات المعلقة اليوم
    pending_today = today_assignments.filter(supervisor_confirmed=False).count()
    
    # آخر 10 تبديلات
    recent_assignments = EmployeeAssignment.objects.filter(
        employee=employee
    ).select_related('sonar', 'shift').order_by('-assigned_at')[:10]
    
    # إحصائيات هذا الأسبوع
    week_start = today - timezone.timedelta(days=today.weekday())
    week_assignments = EmployeeAssignment.objects.filter(
        employee=employee,
        assigned_at__date__gte=week_start
    )
    week_total = week_assignments.count()
    week_confirmed = week_assignments.filter(supervisor_confirmed=True).count()
    
    context = {
        'user_role': 'employee',
        'employee': employee,
        'today_assignments': today_assignments,
        'sonars_today': sonars_today,
        'total_shifts_today': total_shifts_today,
        'confirmed_today': confirmed_today,
        'pending_today': pending_today,
        'recent_assignments': recent_assignments,
        'week_total': week_total,
        'week_confirmed': week_confirmed,
    }
    return render(request, 'dashboards/employee.html', context)


# ==================== إدارة الموظفين ====================

@staff_required
def employee_list(request):
    """عرض قائمة الموظفين - مدمجة مع إدارة الحسابات"""
    user_role = get_user_role(request.user)
    
    # فلترة الموظفين حسب الدور
    if user_role == 'supervisor':
        # المشرف يرى فقط موظفي شفته
        try:
            supervisor = request.user.supervisor_profile
            employees = supervisor.get_employees().order_by('name')
            can_manage_accounts = False  # المشرف لا يمكنه إنشاء حسابات
        except:
            employees = Employee.objects.none()
            can_manage_accounts = False
    else:
        # سوبر أدمن والمدير يرون الكل
        employees = Employee.objects.all().order_by('name')
        can_manage_accounts = user_role in ['superadmin', 'manager']
    
    # إضافة معلومات عن الحسابات
    employees_data = []
    accounts_count = 0
    leave_count = 0
    
    for emp in employees:
        has_account = hasattr(emp, 'user') and emp.user is not None
        employees_data.append({
            'employee': emp,
            'has_account': has_account,
        })
        
        if has_account:
            accounts_count += 1
        if emp.is_on_leave:
            leave_count += 1
    
    context = {
        'employees': employees,
        'employees_data': employees_data,
        'user_role': user_role,
        'can_manage_accounts': can_manage_accounts,
        'accounts_count': accounts_count,
        'leave_count': leave_count,
    }
    return render(request, 'employees/list.html', context)

@staff_required
def employee_create(request):
    """إضافة موظف جديد (مع أو بدون حساب)"""
    user_role = get_user_role(request.user)
    can_create_account = user_role in ['manager', 'superadmin']
    
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            # حفظ بيانات الموظف
            employee = form.save(commit=False)
            if can_create_account:
                employee.created_by = request.user
            
            # إذا تم اختيار إنشاء حساب
            create_account = form.cleaned_data.get('create_account')
            if create_account and can_create_account:
                # إنشاء User
                username = form.cleaned_data.get('username')
                password = form.cleaned_data.get('password')
                
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=employee.name
                )
                
                employee.user = user
                employee.save()
                
                messages.success(
                    request, 
                    f'✅ تم إضافة الموظف {employee.name} مع حساب ({username}) بنجاح!'
                )
            else:
                employee.save()
                messages.success(request, f'✅ تم إضافة الموظف {employee.name} بنجاح!')
            
            return redirect('employee_list')
    else:
        form = EmployeeForm()
    
    context = {
        'form': form, 
        'title': 'إضافة موظف جديد',
        'user_role': user_role,
        'can_create_account': can_create_account
    }
    return render(request, 'employees/form.html', context)

@staff_required
def employee_update(request, pk):
    """تعديل موظف"""
    employee = get_object_or_404(Employee, pk=pk)
    user_role = get_user_role(request.user)
    can_create_account = user_role in ['manager', 'superadmin']
    
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            # حفظ بيانات الموظف
            updated_employee = form.save(commit=False)
            
            # إذا تم اختيار إنشاء حساب ولم يكن لديه حساب مسبقاً
            create_account = form.cleaned_data.get('create_account')
            if create_account and can_create_account and not employee.user:
                # إنشاء User جديد
                username = form.cleaned_data.get('username')
                password = form.cleaned_data.get('password')
                
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=updated_employee.name
                )
                
                updated_employee.user = user
                messages.success(
                    request, 
                    f'✅ تم تحديث بيانات {employee.name} وإنشاء حساب ({username}) بنجاح!'
                )
            else:
                messages.success(request, f'✅ تم تحديث بيانات {employee.name} بنجاح!')
            
            updated_employee.save()
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    
    context = {
        'form': form, 
        'title': f'تعديل: {employee.name}',
        'user_role': user_role,
        'can_create_account': can_create_account,
        'employee': employee
    }
    return render(request, 'employees/form.html', context)

@staff_required
def employee_delete(request, pk):
    """حذف موظف"""
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        name = employee.name
        employee.delete()
        messages.success(request, f'✅ تم حذف الموظف {name} بنجاح!')
        return redirect('employee_list')
    return render(request, 'employees/delete.html', {'employee': employee})


# ==================== إدارة السونارات ====================

@supervisor_required
def sonar_list(request):
    """عرض قائمة السونارات"""
    sonars = Sonar.objects.all().order_by('name')
    return render(request, 'sonars/list.html', {'sonars': sonars})

@supervisor_required
def sonar_create(request):
    """إضافة سونار جديد"""
    if request.method == 'POST':
        form = SonarForm(request.POST)
        if form.is_valid():
            sonar = form.save()
            messages.success(request, f'✅ تم إضافة السونار {sonar.name} بنجاح!')
            return redirect('sonar_list')
    else:
        form = SonarForm()
    return render(request, 'sonars/form.html', {'form': form, 'title': 'إضافة سونار جديد'})

@supervisor_required
def sonar_update(request, pk):
    """تعديل سونار"""
    sonar = get_object_or_404(Sonar, pk=pk)
    if request.method == 'POST':
        form = SonarForm(request.POST, instance=sonar)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ تم تحديث السونار {sonar.name} بنجاح!')
            return redirect('sonar_list')
    else:
        form = SonarForm(instance=sonar)
    return render(request, 'sonars/form.html', {'form': form, 'title': f'تعديل: {sonar.name}'})

@supervisor_required
def sonar_delete(request, pk):
    """حذف سونار"""
    sonar = get_object_or_404(Sonar, pk=pk)
    if request.method == 'POST':
        name = sonar.name
        sonar.delete()
        messages.success(request, f'✅ تم حذف السونار {name} بنجاح!')
        return redirect('sonar_list')
    return render(request, 'sonars/delete.html', {'sonar': sonar})


# ==================== إدارة الورديات ====================

@login_required
def shift_list(request):
    """عرض قائمة الورديات"""
    shifts = Shift.objects.all().order_by('start_hour')
    return render(request, 'shifts/list.html', {'shifts': shifts})

@login_required
def shift_create(request):
    """إضافة شفت جديد"""
    if request.method == 'POST':
        form = ShiftForm(request.POST)
        if form.is_valid():
            shift = form.save()
            messages.success(request, f'✅ تم إضافة الشفت بنجاح!')
            return redirect('shift_list')
    else:
        form = ShiftForm()
    return render(request, 'shifts/form.html', {'form': form, 'title': 'إضافة شفت جديد'})

@login_required
def shift_update(request, pk):
    """تعديل شفت"""
    shift = get_object_or_404(Shift, pk=pk)
    if request.method == 'POST':
        form = ShiftForm(request.POST, instance=shift)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ تم تحديث الشفت بنجاح!')
            return redirect('shift_list')
    else:
        form = ShiftForm(instance=shift)
    return render(request, 'shifts/form.html', {'form': form, 'title': f'تعديل: {shift}'})

@login_required
def shift_delete(request, pk):
    """حذف شفت"""
    shift = get_object_or_404(Shift, pk=pk)
    if request.method == 'POST':
        shift.delete()
        messages.success(request, f'✅ تم حذف الشفت بنجاح!')
        return redirect('shift_list')
    return render(request, 'shifts/delete.html', {'shift': shift})


# ==================== إدارة الجدولة الأسبوعية ====================

@login_required
def weekly_schedule_list(request):
    """عرض قائمة الجدولة الأسبوعية"""
    schedules = WeeklyShiftAssignment.objects.all().select_related('shift').prefetch_related('employees').order_by('-week_start_date')
    return render(request, 'weekly_schedules/list.html', {'schedules': schedules})

@login_required
def weekly_schedule_create(request):
    """إضافة جدولة أسبوعية جديدة"""
    if request.method == 'POST':
        form = WeeklyShiftAssignmentForm(request.POST)
        if form.is_valid():
            schedule = form.save()
            messages.success(request, f'✅ تم إضافة جدولة {schedule.shift} للأسبوع {schedule.week_start_date} بنجاح!')
            return redirect('weekly_schedule_list')
    else:
        form = WeeklyShiftAssignmentForm()
    return render(request, 'weekly_schedules/form.html', {'form': form, 'title': 'إضافة جدولة أسبوعية جديدة'})

@login_required
def weekly_schedule_update(request, pk):
    """تعديل جدولة أسبوعية"""
    schedule = get_object_or_404(WeeklyShiftAssignment, pk=pk)
    if request.method == 'POST':
        form = WeeklyShiftAssignmentForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ تم تحديث الجدولة بنجاح!')
            return redirect('weekly_schedule_list')
    else:
        form = WeeklyShiftAssignmentForm(instance=schedule)
    return render(request, 'weekly_schedules/form.html', {'form': form, 'title': f'تعديل: {schedule.shift}'})

@login_required
def weekly_schedule_delete(request, pk):
    """حذف جدولة أسبوعية"""
    schedule = get_object_or_404(WeeklyShiftAssignment, pk=pk)
    if request.method == 'POST':
        schedule.delete()
        messages.success(request, f'✅ تم حذف الجدولة بنجاح!')
        return redirect('weekly_schedule_list')
    return render(request, 'weekly_schedules/delete.html', {'schedule': schedule})


# ==================== إدارة التبديلات المعلقة (Pending Assignments) ====================

@login_required
def pending_assignments_list(request):
    """عرض قائمة التبديلات المعلقة (غير المؤكدة)"""
    # التبديلات المعلقة (غير مؤكدة ولا يوجد لها confirmation)
    pending = EmployeeAssignment.objects.filter(
        confirmed=False
    ).exclude(
        confirmation__isnull=False
    ).select_related('employee', 'sonar', 'shift').order_by('-assigned_at')
    
    # حساب الإحصائيات
    total_pending = pending.count()
    confirmed_today = AssignmentConfirmation.objects.filter(
        confirmed_at__date=timezone.now().date(),
        status='confirmed'
    ).count()
    rejected_today = AssignmentConfirmation.objects.filter(
        confirmed_at__date=timezone.now().date(),
        status='rejected'
    ).count()
    
    context = {
        'pending_assignments': pending,
        'total_pending': total_pending,
        'confirmed_today': confirmed_today,
        'rejected_today': rejected_today,
    }
    return render(request, 'pending_assignments/list.html', context)

@login_required
def confirm_assignment(request, pk):
    """تأكيد تبديل معلق"""
    assignment = get_object_or_404(EmployeeAssignment, pk=pk, confirmed=False)
    
    # التحقق من أنه لم يتم تأكيده/رفضه مسبقاً
    if hasattr(assignment, 'confirmation'):
        messages.warning(request, '⚠️ هذا التبديل تم معالجته مسبقاً!')
        return redirect('pending_assignments_list')
    
    if request.method == 'POST':
        notes = request.POST.get('notes', '')
        
        # تحديث حالة التبديل
        assignment.confirmed = True
        assignment.save()
        
        # إنشاء سجل تأكيد
        confirmation = AssignmentConfirmation.objects.create(
            assignment=assignment,
            status='confirmed',
            confirmed_by=request.user,
            notes=notes
        )
        
        # إرسال إشعار Telegram للموظف
        if assignment.employee.telegram_id:
            shift_name_ar = dict(assignment.shift.SHIFT_CHOICES).get(
                assignment.shift.name, 
                assignment.shift.name
            )
            
            msg = (
                f"✅ تم تأكيد تبديلك!\n\n"
                f"📢 السونار الجديد: {assignment.sonar.name}\n"
                f"🕒 الشفت: {shift_name_ar}\n"
                f"⏰ الوقت: {timezone.localtime(assignment.assigned_at).strftime('%Y-%m-%d %H:%M')}\n"
                f"👤 تم التأكيد بواسطة: {request.user.username}"
            )
            
            if notes:
                msg += f"\n📝 ملاحظات: {notes}"
            
            send_telegram_message(assignment.employee.telegram_id, msg)
        
        messages.success(
            request, 
            f'✅ تم تأكيد تبديل {assignment.employee.name} إلى {assignment.sonar.name} بنجاح!'
        )
        return redirect('pending_assignments_list')
    
    return render(request, 'pending_assignments/confirm.html', {'assignment': assignment})

@login_required
def reject_assignment(request, pk):
    """رفض تبديل معلق"""
    assignment = get_object_or_404(EmployeeAssignment, pk=pk, confirmed=False)
    
    # التحقق من أنه لم يتم تأكيده/رفضه مسبقاً
    if hasattr(assignment, 'confirmation'):
        messages.warning(request, '⚠️ هذا التبديل تم معالجته مسبقاً!')
        return redirect('pending_assignments_list')
    
    if request.method == 'POST':
        notes = request.POST.get('notes', '')
        employee_name = assignment.employee.name
        sonar_name = assignment.sonar.name
        
        # إنشاء سجل رفض قبل الحذف
        AssignmentConfirmation.objects.create(
            assignment=assignment,
            status='rejected',
            confirmed_by=request.user,
            notes=notes
        )
        
        # إرسال إشعار Telegram للموظف
        if assignment.employee.telegram_id:
            shift_name_ar = dict(assignment.shift.SHIFT_CHOICES).get(
                assignment.shift.name, 
                assignment.shift.name
            )
            
            msg = (
                f"❌ تم رفض تبديلك\n\n"
                f"📢 السونار: {sonar_name}\n"
                f"🕒 الشفت: {shift_name_ar}\n"
                f"⏰ الوقت: {timezone.localtime(assignment.assigned_at).strftime('%Y-%m-%d %H:%M')}\n"
                f"👤 تم الرفض بواسطة: {request.user.username}"
            )
            
            if notes:
                msg += f"\n📝 سبب الرفض: {notes}"
            
            send_telegram_message(assignment.employee.telegram_id, msg)
        
        messages.warning(
            request, 
            f'❌ تم رفض تبديل {employee_name} إلى {sonar_name}'
        )
        return redirect('pending_assignments_list')
    
    return render(request, 'pending_assignments/reject.html', {'assignment': assignment})

@login_required
def confirmed_assignments_list(request):
    """عرض قائمة التبديلات المؤكدة"""
    confirmed = AssignmentConfirmation.objects.filter(
        status='confirmed'
    ).select_related(
        'assignment__employee',
        'assignment__sonar',
        'assignment__shift',
        'confirmed_by'
    ).order_by('-confirmed_at')
    
    return render(request, 'pending_assignments/confirmed_list.html', {'confirmed_assignments': confirmed})

@login_required
def rejected_assignments_list(request):
    """عرض قائمة التبديلات المرفوضة"""
    rejected = AssignmentConfirmation.objects.filter(
        status='rejected'
    ).select_related(
        'assignment__employee',
        'assignment__sonar',
        'assignment__shift',
        'confirmed_by'
    ).order_by('-confirmed_at')
    
    return render(request, 'pending_assignments/rejected_list.html', {'rejected_assignments': rejected})

@login_required
def bulk_confirm_assignments(request):
    """تأكيد جميع التبديلات المعلقة دفعة واحدة"""
    if request.method == 'POST':
        pending = EmployeeAssignment.objects.filter(
            confirmed=False
        ).exclude(
            confirmation__isnull=False
        )
        count = 0
        
        for assignment in pending:
            # تحديث حالة التبديل
            assignment.confirmed = True
            assignment.save()
            
            # إنشاء سجل تأكيد
            AssignmentConfirmation.objects.create(
                assignment=assignment,
                status='confirmed',
                confirmed_by=request.user,
                notes='تأكيد جماعي'
            )
            
            # إرسال إشعار Telegram
            if assignment.employee.telegram_id:
                shift_name_ar = dict(assignment.shift.SHIFT_CHOICES).get(
                    assignment.shift.name, 
                    assignment.shift.name
                )
                
                msg = (
                    f"✅ تم تأكيد تبديلك!\n\n"
                    f"📢 السونار الجديد: {assignment.sonar.name}\n"
                    f"🕒 الشفت: {shift_name_ar}\n"
                    f"⏰ الوقت: {timezone.localtime(assignment.assigned_at).strftime('%Y-%m-%d %H:%M')}"
                )
                
                send_telegram_message(assignment.employee.telegram_id, msg)
            
            count += 1
        
        messages.success(request, f'✅ تم تأكيد {count} تبديل بنجاح!')
        return redirect('pending_assignments_list')
    
    return redirect('pending_assignments_list')


# ==================== التقارير (Reports) ====================

@login_required
def reports_view(request):
    """صفحة التقارير مع فلترة حسب التاريخ والشفت"""
    from datetime import datetime, timedelta, date
    from django.db.models import Count, Q
    
    # الحصول على المعاملات من GET request
    shift_filter = request.GET.get('shift', '')
    status_filter = request.GET.get('status', '')
    
    # إذا لم يتم تحديد تاريخ، استخدم تاريخ اليوم
    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))
    
    # البدء بجميع التبديلات
    assignments = EmployeeAssignment.objects.all().select_related(
        'employee', 'sonar', 'shift', 'supervisor_confirmed_by'
    ).order_by('-assigned_at')
    
    # تطبيق الفلاتر
    if shift_filter:
        assignments = assignments.filter(shift__name=shift_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            assignments = assignments.filter(assigned_at__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            # إضافة يوم كامل لتشمل نهاية اليوم
            date_to_obj = date_to_obj + timedelta(days=1)
            assignments = assignments.filter(assigned_at__lt=date_to_obj)
        except ValueError:
            pass
    
    # تحديث فلتر الحالة ليأخذ في الاعتبار التأكيد الثنائي
    if status_filter == 'confirmed':
        # مؤكد بالكامل (من الموظف والمشرف)
        assignments = assignments.filter(
            employee_confirmed=True,
            supervisor_confirmed=True
        )
    elif status_filter == 'pending':
        # معلق = لم يؤكد من الموظف أو المشرف
        assignments = assignments.filter(
            Q(employee_confirmed=False) | Q(supervisor_confirmed=False)
        )
    elif status_filter == 'waiting_employee':
        # بانتظار تأكيد الموظف
        assignments = assignments.filter(employee_confirmed=False)
    elif status_filter == 'waiting_supervisor':
        # بانتظار تأكيد المشرف (الموظف أكد)
        assignments = assignments.filter(
            employee_confirmed=True,
            supervisor_confirmed=False
        )
    
    # إحصائيات مرتبطة بنتائج البحث
    filtered_assignments = assignments
    total_count = filtered_assignments.count()
    total_confirmed = filtered_assignments.filter(
        employee_confirmed=True,
        supervisor_confirmed=True
    ).count()
    total_pending = filtered_assignments.filter(
        Q(employee_confirmed=False) | Q(supervisor_confirmed=False)
    ).count()
    waiting_employee = filtered_assignments.filter(employee_confirmed=False).count()
    waiting_supervisor = filtered_assignments.filter(
        employee_confirmed=True,
        supervisor_confirmed=False
    ).count()
    
    # إحصائيات حسب الشفت (مرتبطة بالبحث)
    shifts_stats = []
    for shift in Shift.objects.all():
        shift_assignments = filtered_assignments.filter(shift=shift)
        total = shift_assignments.count()
        confirmed = shift_assignments.filter(
            employee_confirmed=True,
            supervisor_confirmed=True
        ).count()
        pending = shift_assignments.filter(
            Q(employee_confirmed=False) | Q(supervisor_confirmed=False)
        ).count()
        
        shifts_stats.append({
            'shift__name': shift.get_name_display(),
            'total': total,
            'confirmed': confirmed,
            'pending': pending
        })
    
    # إحصائيات حسب الموظف (مرتبطة بالبحث) - الأكثر نشاطاً
    employees_stats = filtered_assignments.values(
        'employee__name', 'employee__id'
    ).annotate(
        total=Count('id'),
        confirmed=Count('id', filter=Q(
            employee_confirmed=True,
            supervisor_confirmed=True
        ))
    ).order_by('-total')[:10]
    
    # إحصائيات حسب السونار (مرتبطة بالبحث)
    sonars_stats = filtered_assignments.values(
        'sonar__name'
    ).annotate(
        total=Count('id'),
        confirmed=Count('id', filter=Q(
            employee_confirmed=True,
            supervisor_confirmed=True
        ))
    ).order_by('-total')[:10]
    
    # قائمة جميع الشفتات
    all_shifts = Shift.objects.all()
    
    # 📊 إحصائيات ساعات عمل الموظفين (جميع الموظفين)
    all_employees = Employee.objects.filter(is_on_leave=False).order_by('name')
    employees_work_hours = []
    
    # حساب متوسط ساعات العمل
    total_hours = sum(emp.total_work_hours for emp in all_employees)
    avg_hours = total_hours / all_employees.count() if all_employees.count() > 0 else 0.0
    
    for emp in all_employees:
        diff = emp.total_work_hours - avg_hours
        employees_work_hours.append({
            'name': emp.name,
            'total_work_hours': emp.total_work_hours,
            'diff_from_avg': diff,
            'last_work': emp.last_work_datetime,
            'consecutive_rest': emp.consecutive_rest_count,
            'status': '🔻 فوق المتوسط' if diff > 1.0 else ('🔺 تحت المتوسط' if diff < -1.0 else '⚖️ متوازن')
        })
    
    # ترتيب حسب ساعات العمل (من الأكثر إلى الأقل)
    employees_work_hours.sort(key=lambda x: x['total_work_hours'], reverse=True)
    
    # 📅 سجلات التصفير الشهرية
    from .models import MonthlyWorkHoursReset
    monthly_resets = MonthlyWorkHoursReset.objects.all().order_by('-year', '-month')[:12]  # آخر 12 شهر
    
    context = {
        'assignments': assignments[:100],  # حد أقصى 100 سجل للعرض
        'total_count': total_count,
        'total_confirmed': total_confirmed,
        'total_pending': total_pending,
        'waiting_employee': waiting_employee,
        'waiting_supervisor': waiting_supervisor,
        'shifts_stats': shifts_stats,
        'employees_stats': employees_stats,
        'sonars_stats': sonars_stats,
        'all_shifts': all_shifts,
        'shift_filter': shift_filter,
        'date_from': date_from,
        'date_to': date_to,
        'status_filter': status_filter,
        # إحصائيات ساعات العمل الجديدة
        'employees_work_hours': employees_work_hours,
        'avg_work_hours': avg_hours,
        # سجلات التصفير الشهرية
        'monthly_resets': monthly_resets,
    }
    
    return render(request, 'reports/index.html', context)


@login_required
def employee_performance_report(request):
    """تقرير أداء الموظفين المفصل مع إمكانية التصدير إلى Excel"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    from django.db.models import Count, Q, Sum
    
    # الحصول على الفلاتر
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    export_excel = request.GET.get('export', '') == 'excel'
    
    # جلب جميع الموظفين
    all_employees = Employee.objects.all().order_by('name')
    
    # حساب متوسط ساعات العمل
    total_hours = sum(emp.total_work_hours for emp in all_employees)
    avg_hours = total_hours / all_employees.count() if all_employees.count() > 0 else 0.0
    
    # بناء تقرير مفصل لكل موظف
    employees_data = []
    
    for emp in all_employees:
        # التبديلات الأساسية
        assignments = EmployeeAssignment.objects.filter(employee=emp)
        
        # تطبيق فلتر التاريخ
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                assignments = assignments.filter(assigned_at__gte=date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                date_to_obj = date_to_obj + timedelta(days=1)
                assignments = assignments.filter(assigned_at__lt=date_to_obj)
            except ValueError:
                pass
        
        # إحصائيات التبديلات
        total_assignments = assignments.count()
        confirmed_assignments = assignments.filter(
            employee_confirmed=True,
            supervisor_confirmed=True
        ).count()
        pending_assignments = assignments.filter(
            Q(employee_confirmed=False) | Q(supervisor_confirmed=False)
        ).count()
        
        # حساب ساعات العمل من التبديلات الفعلية (المؤكدة)
        confirmed_work_hours = 0.0
        for assignment in assignments.filter(employee_confirmed=True, supervisor_confirmed=True):
            if assignment.work_duration_hours:
                confirmed_work_hours += assignment.work_duration_hours
        
        # معدل التأكيد
        confirmation_rate = (confirmed_assignments / total_assignments * 100) if total_assignments > 0 else 0.0
        
        # الفرق عن المتوسط
        diff_from_avg = emp.total_work_hours - avg_hours
        
        # الحالة
        if emp.is_on_leave:
            status = '🏖️ في إجازة'
            status_class = 'warning'
        elif diff_from_avg > 5.0:
            status = '🔻 فوق المتوسط بكثير'
            status_class = 'danger'
        elif diff_from_avg > 1.0:
            status = '🔻 فوق المتوسط'
            status_class = 'info'
        elif diff_from_avg < -5.0:
            status = '🔺 تحت المتوسط بكثير'
            status_class = 'success'
        elif diff_from_avg < -1.0:
            status = '🔺 تحت المتوسط'
            status_class = 'primary'
        else:
            status = '⚖️ متوازن'
            status_class = 'secondary'
        
        employees_data.append({
            'name': emp.name,
            'telegram_id': emp.telegram_id or 'غير محدد',
            'total_work_hours': emp.total_work_hours,
            'confirmed_work_hours': confirmed_work_hours,
            'diff_from_avg': diff_from_avg,
            'total_assignments': total_assignments,
            'confirmed_assignments': confirmed_assignments,
            'pending_assignments': pending_assignments,
            'confirmation_rate': confirmation_rate,
            'last_work': emp.last_work_datetime,
            'consecutive_rest': emp.consecutive_rest_count,
            'is_on_leave': emp.is_on_leave,
            'status': status,
            'status_class': status_class,
            'created_at': emp.created_at,
        })
    
    # ترتيب حسب ساعات العمل (من الأكثر إلى الأقل)
    employees_data.sort(key=lambda x: x['total_work_hours'], reverse=True)
    
    # إذا كان طلب تصدير Excel
    if export_excel:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        
        # إنشاء workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'تقرير أداء الموظفين'
        
        # تنسيق العنوان الرئيسي
        ws.merge_cells('A1:M1')
        ws['A1'] = 'تقرير أداء الموظفين - نظام إدارة السونار'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # معلومات التقرير
        ws['A2'] = f'تاريخ الإنشاء: {timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")}'
        ws['A3'] = f'عدد الموظفين: {len(employees_data)}'
        ws['A4'] = f'متوسط ساعات العمل: {avg_hours:.1f} ساعة'
        
        if date_from:
            ws['A5'] = f'من تاريخ: {date_from}'
        if date_to:
            ws['A6'] = f'إلى تاريخ: {date_to}'
        
        # رأس الجدول
        headers = [
            '#', 'اسم الموظف', 'معرف تليجرام', 'إجمالي ساعات العمل',
            'ساعات العمل المؤكدة', 'الفرق عن المتوسط', 'عدد التبديلات',
            'التبديلات المؤكدة', 'التبديلات المعلقة', 'معدل التأكيد %',
            'آخر عمل', 'مرات الراحة', 'الحالة'
        ]
        
        header_row = 8
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF', size=12)
            cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws.row_dimensions[header_row].height = 25
        
        # البيانات
        for idx, emp_data in enumerate(employees_data, 1):
            row = header_row + idx
            
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=emp_data['name'])
            ws.cell(row=row, column=3, value=emp_data['telegram_id'])
            ws.cell(row=row, column=4, value=f"{emp_data['total_work_hours']:.1f}")
            ws.cell(row=row, column=5, value=f"{emp_data['confirmed_work_hours']:.1f}")
            ws.cell(row=row, column=6, value=f"{emp_data['diff_from_avg']:.1f}")
            ws.cell(row=row, column=7, value=emp_data['total_assignments'])
            ws.cell(row=row, column=8, value=emp_data['confirmed_assignments'])
            ws.cell(row=row, column=9, value=emp_data['pending_assignments'])
            ws.cell(row=row, column=10, value=f"{emp_data['confirmation_rate']:.1f}%")
            ws.cell(row=row, column=11, value=emp_data['last_work'].strftime('%Y-%m-%d %H:%M') if emp_data['last_work'] else 'لم يعمل بعد')
            ws.cell(row=row, column=12, value=emp_data['consecutive_rest'])
            ws.cell(row=row, column=13, value=emp_data['status'])
            
            # تلوين الصفوف حسب الحالة
            fill_color = None
            if emp_data['is_on_leave']:
                fill_color = 'FEF3C7'
            elif emp_data['diff_from_avg'] > 5.0:
                fill_color = 'FEE2E2'
            elif emp_data['diff_from_avg'] < -5.0:
                fill_color = 'D1FAE5'
            
            for col in range(1, 14):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        
        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 20
        
        # إنشاء الاستجابة
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'employee_performance_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response
    
    # عرض الصفحة العادية
    context = {
        'employees_data': employees_data,
        'avg_hours': avg_hours,
        'total_employees': len(employees_data),
        'date_from': date_from,
        'date_to': date_to,
        'total_hours': total_hours,
    }
    
    return render(request, 'reports/employee_performance.html', context)


# ==================== إعدادات النظام (System Settings) ====================

@manager_required
def settings_view(request):
    """صفحة إعدادات النظام"""
    # الحصول على الإعدادات الحالية أو إنشاء إعدادات افتراضية
    settings = SystemSettings.get_current_settings()
    
    if request.method == 'POST':
        form = SystemSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            settings_obj = form.save(commit=False)
            settings_obj.updated_by = request.user
            settings_obj.save()
            
            # تحديث جدولة Celery
            update_celery_schedule()
            
            messages.success(request, 'تم حفظ الإعدادات بنجاح!')
            return redirect('settings')
    else:
        form = SystemSettingsForm(instance=settings)
    
    # حساب المعلومات الإضافية
    effective_hours = settings.get_effective_rotation_hours()
    next_rotation = settings.get_next_rotation_time()
    
    context = {
        'form': form,
        'settings': settings,
        'effective_hours': effective_hours,
        'next_rotation': next_rotation,
    }
    
    return render(request, 'settings/index.html', context)


@manager_required
def settings_update(request):
    """تحديث إعدادات النظام"""
    settings = SystemSettings.get_current_settings()
    
    if request.method == 'POST':
        form = SystemSettingsForm(request.POST, instance=settings)
        if form.is_valid():
            settings_obj = form.save(commit=False)
            settings_obj.updated_by = request.user
            settings_obj.save()
            
            # تحديث جدولة Celery
            update_celery_schedule()
            
            messages.success(request, 'تم تحديث الإعدادات بنجاح!')
            return redirect('settings')
        else:
            messages.error(request, 'حدث خطأ في البيانات المدخلة')
    else:
        form = SystemSettingsForm(instance=settings)
    
    return render(request, 'settings/update.html', {'form': form, 'settings': settings})


def update_celery_schedule():
    """تحديث جدولة Celery حسب الإعدادات الحالية"""
    from celery import current_app
    from celery.schedules import crontab
    from datetime import timedelta
    
    settings = SystemSettings.get_current_settings()
    
    # حفظ الجدولة الثابتة للإشعارات المبكرة
    base_schedule = {
        'check-early-notifications': {
            'task': 'shifts.tasks.check_early_notifications_task',
            'schedule': crontab(minute='0,10,20,30,40,50'),  # بداية كل 10 دقائق
        },
    }
    
    if not settings.is_rotation_active:
        # فقط الإشعارات المبكرة بدون التبديل
        current_app.conf.beat_schedule = base_schedule
        print("🔕 تم إيقاف جدولة التبديل التلقائي (الإشعارات المبكرة لا تزال نشطة)")
        return
    
    # حساب الفترة بالساعات
    rotation_hours = float(settings.get_effective_rotation_hours())
    
    # إضافة جدولة التبديل الديناميكية باستخدام timedelta
    base_schedule['rotate-shifts-dynamic'] = {
        'task': 'shifts.tasks.rotate_shifts_task',
        'schedule': timedelta(hours=rotation_hours),  # استخدام timedelta للمرونة
        'args': ()  # سيستخدم الإعدادات المحفوظة
    }
    
    current_app.conf.beat_schedule = base_schedule
    print(f"⏰ تم تحديث جدولة التبديل: كل {rotation_hours} ساعة")


# ==================== تصدير التقارير (Export Reports) ====================

@login_required
def export_reports_excel(request):
    """تصدير التقارير إلى Excel"""
    from datetime import datetime, timedelta, date
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from django.db.models import Q
    
    # الحصول على نفس الفلاتر من صفحة التقارير
    shift_filter = request.GET.get('shift', '')
    status_filter = request.GET.get('status', '')
    
    # إذا لم يتم تحديد تاريخ، استخدم تاريخ اليوم
    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))
    
    # البدء بجميع التبديلات
    assignments = EmployeeAssignment.objects.all().select_related(
        'employee', 'sonar', 'shift', 'supervisor_confirmed_by'
    ).order_by('-assigned_at')
    
    # تطبيق الفلاتر
    if shift_filter:
        assignments = assignments.filter(shift__name=shift_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            assignments = assignments.filter(assigned_at__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            assignments = assignments.filter(assigned_at__lt=date_to_obj)
        except ValueError:
            pass
    
    # تحديث فلتر الحالة ليأخذ في الاعتبار التأكيد الثنائي
    if status_filter == 'confirmed':
        # مؤكد بالكامل (من الموظف والمشرف)
        assignments = assignments.filter(
            employee_confirmed=True,
            supervisor_confirmed=True
        )
    elif status_filter == 'pending':
        # معلق = لم يؤكد من الموظف أو المشرف
        assignments = assignments.filter(
            Q(employee_confirmed=False) | Q(supervisor_confirmed=False)
        )
    elif status_filter == 'waiting_employee':
        # بانتظار تأكيد الموظف
        assignments = assignments.filter(employee_confirmed=False)
    elif status_filter == 'waiting_supervisor':
        # بانتظار تأكيد المشرف (الموظف أكد)
        assignments = assignments.filter(
            employee_confirmed=True,
            supervisor_confirmed=False
        )
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير التبديلات"
    
    # تعريف الأنماط
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # إضافة العنوان
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = "تقرير التبديلات - نظام إدارة السونار"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # إضافة معلومات الفلتر
    row = 3
    if shift_filter or date_from or date_to or status_filter:
        ws.merge_cells(f'A{row}:J{row}')
        filter_cell = ws[f'A{row}']
        filter_info = "الفلاتر المطبقة: "
        if shift_filter:
            filter_info += f"الشفت: {shift_filter} | "
        if date_from:
            filter_info += f"من: {date_from} | "
        if date_to:
            filter_info += f"إلى: {date_to} | "
        if status_filter:
            status_map = {
                'confirmed': 'مؤكد كلياً',
                'pending': 'معلق',
                'waiting_employee': 'بانتظار الموظف',
                'waiting_supervisor': 'بانتظار المشرف'
            }
            status_text = status_map.get(status_filter, status_filter)
            filter_info += f"الحالة: {status_text}"
        filter_cell.value = filter_info
        filter_cell.alignment = Alignment(horizontal="center")
        row += 2
    
    # إضافة رأس الجدول
    headers = ['#', 'الموظف', 'السونار', 'الشفت', 'التاريخ', 'الوقت', 'تأكيد الموظف', 'تأكيد المشرف', 'المشرف', 'الحالة']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # إضافة البيانات
    row += 1
    for idx, assignment in enumerate(assignments, 1):
        # التحقق من وجود البيانات الأساسية
        if not assignment.employee:
            continue  # تخطي التبديلات بدون موظف
        
        # ترجمة اسم الشفت
        shift_name_mapping = {
            'morning': 'صباحي',
            'evening': 'مسائي',
            'night': 'ليلي'
        }
        shift_name_ar = '-'
        if assignment.shift:
            shift_name_ar = shift_name_mapping.get(assignment.shift.name, assignment.shift.name)
        
        # وقت تأكيد الموظف
        employee_confirmed_time = ''
        if assignment.employee_confirmed_at:
            employee_confirmed_time = timezone.localtime(assignment.employee_confirmed_at).strftime('%Y-%m-%d %H:%M')
        
        # وقت تأكيد المشرف
        supervisor_confirmed_time = ''
        if assignment.supervisor_confirmed_at:
            supervisor_confirmed_time = timezone.localtime(assignment.supervisor_confirmed_at).strftime('%Y-%m-%d %H:%M')
        
        # اسم المشرف
        supervisor_name = ''
        if assignment.supervisor_confirmed_by:
            supervisor_name = assignment.supervisor_confirmed_by.first_name or assignment.supervisor_confirmed_by.username
        
        # الحالة النهائية
        if assignment.supervisor_confirmed:
            status = "مؤكد كلياً ✅✅"
        elif assignment.employee_confirmed:
            status = "بانتظار المشرف ⏳"
        else:
            status = "بانتظار الموظف ⏰"
        
        # اسم الموظف
        employee_name = assignment.employee.name if assignment.employee.name else 'غير محدد'
        
        # اسم السونار (قد يكون None للموظفين في الاحتياط)
        sonar_name = 'احتياط'
        if assignment.sonar and assignment.sonar.name:
            sonar_name = assignment.sonar.name
        
        # كتابة البيانات
        data = [
            idx,
            employee_name,
            sonar_name,
            shift_name_ar,
            timezone.localtime(assignment.assigned_at).strftime('%Y-%m-%d'),
            timezone.localtime(assignment.assigned_at).strftime('%H:%M'),
            employee_confirmed_time if employee_confirmed_time else 'لم يؤكد',
            supervisor_confirmed_time if supervisor_confirmed_time else 'لم يؤكد',
            supervisor_name if supervisor_name else '-',
            status
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
            # تلوين الحالة
            if col_num == 10:  # عمود الحالة
                if assignment.supervisor_confirmed:
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                elif assignment.employee_confirmed:
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        
        row += 1
    
    # إضافة الإحصائيات في النهاية
    row += 2
    total_count = assignments.count()
    confirmed_count = assignments.filter(supervisor_confirmed=True).count()
    pending_count = assignments.filter(supervisor_confirmed=False).count()
    
    ws.merge_cells(f'A{row}:B{row}')
    stats_cell = ws[f'A{row}']
    stats_cell.value = f"إجمالي التبديلات: {total_count}"
    stats_cell.font = Font(bold=True)
    
    ws.merge_cells(f'C{row}:D{row}')
    confirmed_cell = ws[f'C{row}']
    confirmed_cell.value = f"المؤكدة: {confirmed_count}"
    confirmed_cell.font = Font(bold=True, color="10B981")
    
    ws.merge_cells(f'E{row}:F{row}')
    pending_cell = ws[f'E{row}']
    pending_cell.value = f"المعلقة: {pending_count}"
    pending_cell.font = Font(bold=True, color="F59E0B")
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 6   # #
    ws.column_dimensions['B'].width = 20  # الموظف
    ws.column_dimensions['C'].width = 18  # السونار
    ws.column_dimensions['D'].width = 12  # الشفت
    ws.column_dimensions['E'].width = 12  # التاريخ
    ws.column_dimensions['F'].width = 10  # الوقت
    ws.column_dimensions['G'].width = 18  # تأكيد الموظف
    ws.column_dimensions['H'].width = 18  # تأكيد المشرف
    ws.column_dimensions['I'].width = 15  # المشرف
    ws.column_dimensions['J'].width = 18  # الحالة
    
    # إعداد الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'تقرير_التبديلات_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_reports_pdf(request):
    """تصدير التقارير إلى PDF"""
    from datetime import datetime, timedelta, date
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    from django.http import HttpResponse
    from django.db.models import Q
    import arabic_reshaper
    from bidi.algorithm import get_display
    import os
    
    # تسجيل الخط العربي
    try:
        # محاولة استخدام خط Arial (متوفر في Windows)
        arial_path = 'C:/Windows/Fonts/arial.ttf'
        arial_bold_path = 'C:/Windows/Fonts/arialbd.ttf'
        
        if os.path.exists(arial_path):
            pdfmetrics.registerFont(TTFont('Arabic', arial_path))
            if os.path.exists(arial_bold_path):
                pdfmetrics.registerFont(TTFont('Arabic-Bold', arial_bold_path))
            else:
                pdfmetrics.registerFont(TTFont('Arabic-Bold', arial_path))
        else:
            # استخدام خط احتياطي
            pdfmetrics.registerFont(TTFont('Arabic', 'Helvetica'))
            pdfmetrics.registerFont(TTFont('Arabic-Bold', 'Helvetica-Bold'))
    except:
        # في حالة فشل التسجيل، استخدم الخط الافتراضي
        pdfmetrics.registerFont(TTFont('Arabic', 'Helvetica'))
        pdfmetrics.registerFont(TTFont('Arabic-Bold', 'Helvetica-Bold'))
    
    # دالة لتنسيق النص العربي
    def format_arabic(text):
        if text:
            reshaped_text = arabic_reshaper.reshape(str(text))
            return get_display(reshaped_text)
        return ""
    
    # الحصول على نفس الفلاتر من صفحة التقارير
    shift_filter = request.GET.get('shift', '')
    status_filter = request.GET.get('status', '')
    
    # إذا لم يتم تحديد تاريخ، استخدم تاريخ اليوم
    today = date.today()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))
    
    # البدء بجميع التبديلات
    assignments = EmployeeAssignment.objects.all().select_related(
        'employee', 'sonar', 'shift', 'supervisor_confirmed_by'
    ).order_by('-assigned_at')
    
    # تطبيق الفلاتر
    if shift_filter:
        assignments = assignments.filter(shift__name=shift_filter)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            assignments = assignments.filter(assigned_at__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj + timedelta(days=1)
            assignments = assignments.filter(assigned_at__lt=date_to_obj)
        except ValueError:
            pass
    
    # تحديث فلتر الحالة ليأخذ في الاعتبار التأكيد الثنائي
    if status_filter == 'confirmed':
        # مؤكد بالكامل (من الموظف والمشرف)
        assignments = assignments.filter(
            employee_confirmed=True,
            supervisor_confirmed=True
        )
    elif status_filter == 'pending':
        # معلق = لم يؤكد من الموظف أو المشرف
        assignments = assignments.filter(
            Q(employee_confirmed=False) | Q(supervisor_confirmed=False)
        )
    elif status_filter == 'waiting_employee':
        # بانتظار تأكيد الموظف
        assignments = assignments.filter(employee_confirmed=False)
    elif status_filter == 'waiting_supervisor':
        # بانتظار تأكيد المشرف (الموظف أكد)
        assignments = assignments.filter(
            employee_confirmed=True,
            supervisor_confirmed=False
        )
    
    # إنشاء ملف PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f'تقرير_التبديلات_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # إنشاء المستند
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    
    # العنوان
    title_text = format_arabic("تقرير التبديلات - نظام إدارة السونار")
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=getSampleStyleSheet()['Heading1'],
        alignment=TA_CENTER,
        fontSize=16,
        spaceAfter=20,
        fontName='Arabic-Bold'
    )
    elements.append(Paragraph(title_text, title_style))
    
    # معلومات الفلتر
    if shift_filter or date_from or date_to or status_filter:
        filter_info = format_arabic("الفلاتر المطبقة: ")
        if shift_filter:
            filter_info += format_arabic(f"الشفت: {shift_filter} | ")
        if date_from:
            filter_info += format_arabic(f"من: {date_from} | ")
        if date_to:
            filter_info += format_arabic(f"إلى: {date_to} | ")
        if status_filter:
            status_map = {
                'confirmed': 'مؤكد كلياً',
                'pending': 'معلق',
                'waiting_employee': 'بانتظار الموظف',
                'waiting_supervisor': 'بانتظار المشرف'
            }
            status_text = status_map.get(status_filter, status_filter)
            filter_info += format_arabic(f"الحالة: {status_text}")
        
        filter_style = ParagraphStyle(
            'FilterStyle',
            parent=getSampleStyleSheet()['Normal'],
            alignment=TA_CENTER,
            fontSize=10,
            fontName='Arabic'
        )
        elements.append(Paragraph(filter_info, filter_style))
        elements.append(Spacer(1, 20))
    
    # إعداد بيانات الجدول
    data = []
    
    # رأس الجدول
    headers = [
        '#',
        format_arabic('التاريخ'),
        format_arabic('الوقت'),
        format_arabic('الموظف'),
        format_arabic('السونار'),
        format_arabic('الشفت'),
        format_arabic('تأكيد موظف'),
        format_arabic('المشرف'),
        format_arabic('تأكيد مشرف'),
        format_arabic('الحالة')
    ]
    data.append(headers)
    
    # البيانات
    shift_name_mapping = {
        'morning': 'صباحي',
        'evening': 'مسائي',
        'night': 'ليلي'
    }
    
    for idx, assignment in enumerate(assignments[:100], 1):  # حد أقصى 100 سجل
        # التحقق من وجود البيانات الأساسية
        if not assignment.employee:
            continue  # تخطي التبديلات بدون موظف
        
        # اسم الشفت
        shift_name_ar = '-'
        if assignment.shift:
            shift_name_ar = shift_name_mapping.get(assignment.shift.name, assignment.shift.name)
        
        # الحالة النهائية
        if assignment.supervisor_confirmed:
            status = format_arabic("مؤكد كلياً")
        elif assignment.employee_confirmed:
            status = format_arabic("انتظار مشرف")
        else:
            status = format_arabic("انتظار موظف")
        
        # وقت تأكيد الموظف
        emp_confirm = ''
        if assignment.employee_confirmed_at:
            emp_confirm = timezone.localtime(assignment.employee_confirmed_at).strftime('%m-%d %H:%M')
        else:
            emp_confirm = format_arabic('لم يؤكد')
        
        # وقت تأكيد المشرف
        sup_confirm = ''
        if assignment.supervisor_confirmed_at:
            sup_confirm = timezone.localtime(assignment.supervisor_confirmed_at).strftime('%m-%d %H:%M')
        else:
            sup_confirm = format_arabic('لم يؤكد')
        
        # اسم المشرف
        supervisor_name = '-'
        if assignment.supervisor_confirmed_by:
            supervisor_name = format_arabic(
                assignment.supervisor_confirmed_by.first_name or 
                assignment.supervisor_confirmed_by.username
            )
        
        # اسم السونار (قد يكون None للموظفين في الاحتياط)
        sonar_name = format_arabic('احتياط') if not assignment.sonar else format_arabic(assignment.sonar.name)
        
        # اسم الموظف
        employee_name = format_arabic(assignment.employee.name) if assignment.employee.name else format_arabic('غير محدد')
        
        row = [
            str(idx),
            timezone.localtime(assignment.assigned_at).strftime('%Y-%m-%d'),
            timezone.localtime(assignment.assigned_at).strftime('%H:%M'),
            employee_name,
            sonar_name,
            format_arabic(shift_name_ar),
            emp_confirm,
            supervisor_name,
            sup_confirm,
            status
        ]
        data.append(row)
    
    # إنشاء الجدول
    table = Table(data, repeatRows=1)
    
    # تنسيق الجدول
    table.setStyle(TableStyle([
        # رأس الجدول
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Arabic-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        # البيانات
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
    ]))
    
    elements.append(table)
    
    # الإحصائيات
    elements.append(Spacer(1, 20))
    total_count = assignments.count()
    confirmed_count = assignments.filter(supervisor_confirmed=True).count()
    pending_count = assignments.filter(supervisor_confirmed=False).count()
    
    stats_text = format_arabic(
        f"إجمالي التبديلات: {total_count} | المؤكدة: {confirmed_count} | المعلقة: {pending_count}"
    )
    stats_style = ParagraphStyle(
        'StatsStyle',
        parent=getSampleStyleSheet()['Normal'],
        alignment=TA_CENTER,
        fontSize=12,
        textColor=colors.HexColor('#1F2937'),
        fontName='Arabic-Bold'
    )
    elements.append(Paragraph(stats_text, stats_style))
    
    # بناء المستند
    doc.build(elements)
    return response


# ==================== إدارة الحسابات (Manager & SuperAdmin) ====================

@manager_required
def supervisor_accounts_list(request):
    """قائمة حسابات المشرفين"""
    if request.user.is_superuser:
        supervisors = Supervisor.objects.all().select_related('user').order_by('-created_at')
    else:
        supervisors = Supervisor.objects.filter(created_by=request.user).select_related('user').order_by('-created_at')
    
    context = {
        'supervisors': supervisors,
        'user_role': get_user_role(request.user)
    }
    return render(request, 'accounts/supervisor_list.html', context)


@manager_required
def supervisor_account_create(request):
    """إنشاء حساب مشرف جديد"""
    if request.method == 'POST':
        form = SupervisorCreateForm(request.POST)
        if form.is_valid():
            # إنشاء User
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=form.cleaned_data.get('name')
            )
            
            # إنشاء Supervisor profile
            supervisor = form.save(commit=False)
            supervisor.user = user
            supervisor.created_by = request.user
            supervisor.save()
            
            messages.success(request, f'✅ تم إنشاء حساب المشرف {supervisor.name} بنجاح!')
            return redirect('supervisor_accounts_list')
    else:
        form = SupervisorCreateForm()
    
    context = {
        'form': form,
        'title': 'إضافة مشرف جديد',
        'user_role': get_user_role(request.user)
    }
    return render(request, 'accounts/supervisor_form.html', context)


@manager_required
def supervisor_account_toggle(request, pk):
    """تفعيل/إلغاء تفعيل حساب مشرف"""
    if request.user.is_superuser:
        supervisor = get_object_or_404(Supervisor, pk=pk)
    else:
        supervisor = get_object_or_404(Supervisor, pk=pk, created_by=request.user)
    
    supervisor.is_active = not supervisor.is_active
    supervisor.save()
    
    status = 'تفعيل' if supervisor.is_active else 'إلغاء تفعيل'
    messages.success(request, f'✅ تم {status} حساب المشرف {supervisor.name}')
    return redirect('supervisor_accounts_list')


@manager_required
def supervisor_account_delete(request, pk):
    """حذف حساب مشرف"""
    if request.user.is_superuser:
        supervisor = get_object_or_404(Supervisor, pk=pk)
    else:
        supervisor = get_object_or_404(Supervisor, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        name = supervisor.name
        user = supervisor.user
        supervisor.delete()
        user.delete()
        messages.success(request, f'✅ تم حذف حساب المشرف {name}')
        return redirect('supervisor_accounts_list')
    
    context = {
        'supervisor': supervisor,
        'user_role': get_user_role(request.user)
    }
    return render(request, 'accounts/supervisor_delete.html', context)


@manager_required
def employee_accounts_list(request):
    """قائمة حسابات الموظفين - إعادة توجيه للقائمة الموحدة"""
    return redirect('employee_list')


@manager_required
def employee_account_create(request):
    """إنشاء حساب موظف جديد - إعادة توجيه للنموذج الموحد"""
    # الآن يتم إنشاء الموظفين مع الحسابات من نفس النموذج
    return redirect('employee_create')


@manager_required
def employee_account_delete(request, pk):
    """حذف حساب موظف"""
    if request.user.is_superuser:
        employee = get_object_or_404(Employee, pk=pk, user__isnull=False)
    else:
        employee = get_object_or_404(Employee, pk=pk, created_by=request.user, user__isnull=False)
    
    if request.method == 'POST':
        name = employee.name
        user = employee.user
        employee.delete()
        if user:
            user.delete()
        messages.success(request, f'✅ تم حذف حساب الموظف {name}')
        return redirect('employee_accounts_list')
    
    context = {
        'employee': employee,
        'user_role': get_user_role(request.user)
    }
    return render(request, 'accounts/employee_delete.html', context)


@superadmin_required
def manager_accounts_list(request):
    """قائمة حسابات المديرين (سوبر أدمن فقط)"""
    managers = Manager.objects.all().select_related('user').order_by('-created_at')
    
    context = {
        'managers': managers,
        'user_role': 'superadmin'
    }
    return render(request, 'accounts/manager_list.html', context)


@superadmin_required
def manager_account_create(request):
    """إنشاء حساب مدير جديد (سوبر أدمن فقط)"""
    if request.method == 'POST':
        form = ManagerCreateForm(request.POST)
        if form.is_valid():
            # إنشاء User
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=form.cleaned_data.get('name')
            )
            
            # إنشاء Manager profile
            manager = form.save(commit=False)
            manager.user = user
            manager.save()
            
            messages.success(request, f'✅ تم إنشاء حساب المدير {manager.name} بنجاح!')
            return redirect('manager_accounts_list')
    else:
        form = ManagerCreateForm()
    
    context = {
        'form': form,
        'title': 'إضافة مدير جديد',
        'user_role': 'superadmin'
    }
    return render(request, 'accounts/manager_form.html', context)


@superadmin_required
def manager_account_toggle(request, pk):
    """تفعيل/إلغاء تفعيل حساب مدير"""
    manager = get_object_or_404(Manager, pk=pk)
    
    manager.is_active = not manager.is_active
    manager.save()
    
    status = 'تفعيل' if manager.is_active else 'إلغاء تفعيل'
    messages.success(request, f'✅ تم {status} حساب المدير {manager.name}')
    return redirect('manager_accounts_list')


@superadmin_required
def manager_account_delete(request, pk):
    """حذف حساب مدير"""
    manager = get_object_or_404(Manager, pk=pk)
    
    if request.method == 'POST':
        name = manager.name
        user = manager.user
        manager.delete()
        user.delete()
        messages.success(request, f'✅ تم حذف حساب المدير {name}')
        return redirect('manager_accounts_list')
    
    context = {
        'manager': manager,
        'user_role': 'superadmin'
    }
    return render(request, 'accounts/manager_delete.html', context)


# ==================== نظام التأكيد الثنائي ====================

@employee_required
def employee_confirm_assignment(request, pk):
    """تأكيد الموظف أنه ذهب للسونار"""
    assignment = get_object_or_404(EmployeeAssignment, pk=pk)
    
    # التحقق من أن الموظف يؤكد تبديله الخاص
    try:
        employee = request.user.employee_profile
        if assignment.employee != employee:
            messages.error(request, '⛔ لا يمكنك تأكيد تبديل موظف آخر!')
            return redirect('employee_dashboard')
    except:
        messages.error(request, '⛔ حدث خطأ!')
        return redirect('home')
    
    # تأكيد التبديل
    if not assignment.employee_confirmed:
        assignment.employee_confirmed = True
        assignment.employee_confirmed_at = timezone.now()
        assignment.save()
        messages.success(request, f'✅ تم تأكيد ذهابك للسونار {assignment.sonar.name}')
    else:
        messages.info(request, 'ℹ️ هذا التبديل مؤكد مسبقاً!')
    
    return redirect('employee_dashboard')


@supervisor_required
def supervisor_confirm_assignment(request, pk):
    """تأكيد المشرف على تأكيد الموظف"""
    assignment = get_object_or_404(EmployeeAssignment, pk=pk)
    
    # التحقق من أن التبديل مؤكد من الموظف أولاً
    if not assignment.employee_confirmed:
        messages.warning(request, '⚠️ يجب على الموظف تأكيد الذهاب للسونار أولاً!')
        return redirect('supervisor_dashboard')
    
    # تأكيد المشرف
    if not assignment.supervisor_confirmed:
        assignment.supervisor_confirmed = True
        assignment.supervisor_confirmed_at = timezone.now()
        assignment.supervisor_confirmed_by = request.user
        assignment.confirmed = True  # تأكيد نهائي
        assignment.save()
        messages.success(
            request, 
            f'✅ تم تأكيد تبديل {assignment.employee.name} للسونار {assignment.sonar.name}'
        )
    else:
        messages.info(request, 'ℹ️ هذا التبديل مؤكد مسبقاً!')
    
    return redirect('supervisor_dashboard')


# ==================== الإشعارات المخصصة (Custom Notifications) ====================

@staff_required
def send_custom_notification(request):
    """إرسال إشعار مخصص للموظفين"""
    if request.method == 'POST':
        form = CustomNotificationForm(request.POST)
        if form.is_valid():
            notification = form.save(commit=False)
            notification.sent_by = request.user
            notification.save()
            
            # تحديد الموظفين المستهدفين
            if notification.send_to_all:
                employees = Employee.objects.filter(telegram_id__isnull=False).exclude(telegram_id='')
            else:
                form.save_m2m()  # حفظ العلاقة many-to-many
                employees = notification.target_employees.filter(telegram_id__isnull=False).exclude(telegram_id='')
            
            # إرسال الإشعار لكل موظف
            sent_count = 0
            for employee in employees:
                message = f"""
📢 {notification.title}

{notification.message}

━━━━━━━━━━━━━━━━━
📤 من: {notification.sent_by.get_full_name() or notification.sent_by.username}
📅 {timezone.localtime(notification.sent_at).strftime('%Y-%m-%d %H:%M')}
                """
                
                try:
                    send_telegram_message(employee.telegram_id, message)
                    sent_count += 1
                except Exception as e:
                    print(f"❌ خطأ في إرسال الإشعار لـ {employee.name}: {e}")
            
            # تحديث عدد المرسل إليهم
            notification.total_sent = sent_count
            notification.save()
            
            if sent_count > 0:
                messages.success(
                    request,
                    f'✅ تم إرسال الإشعار بنجاح لـ {sent_count} موظف'
                )
            else:
                messages.warning(request, '⚠️ لم يتم إرسال أي إشعار (تحقق من أرقام التليجرام)')
            
            return redirect('custom_notifications_list')
    else:
        form = CustomNotificationForm()
    
    return render(request, 'notifications/send.html', {'form': form})


@staff_required
def custom_notifications_list(request):
    """عرض قائمة الإشعارات المخصصة المرسلة"""
    notifications = CustomNotification.objects.all().order_by('-sent_at')
    return render(request, 'notifications/list.html', {'notifications': notifications})


@staff_required
def custom_notification_detail(request, pk):
    """تفاصيل إشعار مخصص"""
    notification = get_object_or_404(CustomNotification, pk=pk)
    return render(request, 'notifications/detail.html', {'notification': notification})


@staff_required
def expired_assignments_list(request):
    """عرض قائمة الطلبات المنتهية غير المؤكدة"""
    # الطلبات المنتهية (لم يؤكدها الموظف وفات الوقت)
    expired_assignments = EmployeeAssignment.objects.filter(
        is_expired_unconfirmed=True,
        confirmed=False  # لم يتم تأكيدها نهائياً
    ).select_related('employee', 'sonar', 'shift').order_by('-expired_at')
    
    # حساب الإحصائيات
    total_expired = expired_assignments.count()
    
    # حساب الوقت المنقضي لكل طلب
    now = timezone.now()
    for assignment in expired_assignments:
        if assignment.expired_at:
            time_diff = now - assignment.expired_at
            hours_diff = int(time_diff.total_seconds() / 3600)
            days_diff = hours_diff // 24
            remaining_hours = hours_diff % 24
            
            if days_diff > 0:
                assignment.time_since_expiry = f"{days_diff} يوم و {remaining_hours} ساعة"
            else:
                assignment.time_since_expiry = f"{hours_diff} ساعة"
        else:
            assignment.time_since_expiry = "غير محدد"
    
    context = {
        'expired_assignments': expired_assignments,
        'total_expired': total_expired,
    }
    return render(request, 'pending_assignments/expired_list.html', context)